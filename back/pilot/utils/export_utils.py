import re

import arrow
import tempfile
import zipfile
import logging

from django.db.models import ForeignKey
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter

from django.utils.translation import ugettext_lazy as _
from django.conf import settings
from django.utils.encoding import force_text
from django.template.defaultfilters import yesno
from django.utils.safestring import mark_safe
from django.db.models.fields.related import ManyToManyField
from django.db.models.fields import DateTimeField, DateField, BooleanField

from pilot.notifications.const import NotificationType
from pilot.pilot_users.models import PilotUser
from pilot.queue.jobs import Job
from pilot.queue.rq_setup import low_priority_queue
from pilot.utils.s3 import upload_s3_file_multipart
from pilot.utils.serialization import SerializationDialect

logger = logging.getLogger(__name__)


ASSET_METADATA_FIELDS = (
    'title',
    'description',
    'filetype',
    'extension',
    'mime',
    'version',
    'size',
    # Tracking fields
    'created_by',
    'created_at',
    'updated_by',
    'updated_at',
)
PROJECT_METADATA_FIELDS = (
    'id',
    'name',
    'description',
    'start',
    'end',
    'channels',
    'state',
    'owners',
    'closed_at',
    'targets',
    'priority',
    'category',
    'tags',
    'copied_from_id',
    # Tracking fields
    'created_by',
    'created_at',
    'updated_by',
    'updated_at',
)
ITEM_METADATA_FIELDS = (
    'id',
    'item_type',
    'channels',
    'targets',
    'project',
    'workflow_state',
    'in_trash',
    'language',
    'owners',
    'tags',
    'copied_from_id',
    # Tracking fields
    'created_by',
    'created_at',
    'updated_by',
    'updated_at',
)
USER_METADATA_FIELDS = (
    'username',
    'email',
    'first_name',
    'last_name',
    'phone',
    'localization',
    'job',
    'teams',
    'date_joined',
)


def format_date(date):
    if not date:
        return ''
    return arrow.get(date.isoformat()).format('DD/MM/YYYY')


def format_date_time(date_time):
    if not date_time:
        return ''
    return arrow.get(date_time.isoformat()).format('DD/MM/YYYY HH:MM')


def format_related_instance(instance, raw_id=False):
    # We simply use the raw id
    if raw_id:
        return instance.id
    # We want to avoid the arobase in the standard representation of a user (@username)
    # so excel won't go nut.
    elif isinstance(instance, PilotUser):
        return instance.username
    # str(value) will force the repr of ForeignKey
    else:
        return str(instance)


def format_field(instance, field_name, dialect=SerializationDialect.TEXT):
    field = instance._meta.get_field(field_name)
    value = getattr(instance, field.name)

    if value is None:
        value = ''
    else:
        # str(choice) will force the translation of Promise
        if field.choices:
            value = str(dict(field.choices).get(value))

        elif isinstance(field, ManyToManyField):
            value = dialect.newline.join([format_related_instance(related) for related in value.all()])
            if dialect == SerializationDialect.HTML:
                value = mark_safe(value)

        elif isinstance(field, ForeignKey):
            # If the id field is used, we use the raw_id
            value = format_related_instance(
                instance=value,
                raw_id=field_name.endswith('_id')
            )

        elif isinstance(field, BooleanField):
            value = yesno(value)

        elif isinstance(field, DateField):
            value = format_date(value)

        elif isinstance(field, DateTimeField):
            value = format_date_time(value)

        # Default on the standard field string representation
        else:
            value = str(value)

    return value


class BaseXLSExportJob(Job):
    queue = low_priority_queue
    download_export_string = _("Télécharger l'export")
    email_subject = _("Export terminé")
    notification_message = _("L'export que vous avez demandé est terminé. "
                             "<a href='{export_url}'>Cliquez ici pour le télécharger</a>")

    def run(self):
        self.run_export()

    def run_export(self):
        temp_file = tempfile.NamedTemporaryFile()

        self.write_data_to_xls_file(temp_file)

        s3_key = 'export/{deskId}/{uuid}'.format(
            deskId=self.job_tracker.desk.id,
            uuid=self.job_tracker.job_id
        )
        file_name = self.get_file_name()

        # Rewind the temporary file storage to the beginning
        temp_file.seek(0)
        # Upload the archive to S3
        upload_s3_file_multipart(s3_key, temp_file.file, file_name)
        # Close the temporary file
        temp_file.close()
        # Save the result in the JobTracker
        self.job_tracker.data = {
            'result_file_name': file_name,
            'result_url': settings.AWS_S3_BASE_URL + s3_key
        }
        self.job_tracker.save()
        # Notify the user that the export is finished
        self.notify_xls_export_completed()

    def write_data_to_xls_file(self, temp_file):
        raise NotImplementedError()

    def get_file_name(self):
        raise NotImplementedError()

    def notify_xls_export_completed(self):
        from pilot.notifications.notify import notify

        notify(
            desk=self.job_tracker.desk,
            type=NotificationType.EXPORT_XLS,
            to_users=[self.job_tracker.created_by],
            message=self.notification_message,
            email_subject=self.email_subject,
            target_url=self.job_tracker.data['result_url'],
            button_action_text=self.download_export_string,
            data=self.get_notification_data(),
            context=self.get_notification_context(),
        )

    def get_notification_context(self):
        return {
            'export_url': self.job_tracker.data['result_url'],
        }

    def get_notification_data(self):
        return {
            'export_type': self.job_type
        }


class BaseXLSExporter(object):
    model = None
    metadata_fields = []

    def __init__(self, queryset, output_file):
        self.queryset = queryset
        self.output_file = output_file

        self.workbook = Workbook()
        self.worksheet = self.workbook.active

        # Indicies of the column where the text should wrap (1-based, as openpyxl)
        self.wrapped_columns = []

    def get_header(self):
        for i, metadata_field_name in enumerate(self.metadata_fields):
            field = self.model._meta.get_field(metadata_field_name)
            if isinstance(field, ManyToManyField):
                self.wrapped_columns.append(i+1)
            yield field.verbose_name

    def get_row(self, i, instance):
        for metadata_field_name in self.metadata_fields:
            yield format_field(instance, metadata_field_name)

    def get_all_column_letters(self):
        for col_index in range(0, self.worksheet.max_column):
            yield get_column_letter(col_index + 1)

    def add_style(self):
        # Set a width of 25 on all columns
        for column_letter in self.get_all_column_letters():
            self.worksheet.column_dimensions[column_letter].width = 25

        for row_index in range(2, self.worksheet.max_row + 1):
            for column_index in self.wrapped_columns:
                self.worksheet.cell(row_index, column_index).alignment = Alignment(wrap_text=True)

    def clean_illegal_characters(self, row):
        ILLEGAL_CHARACTERS = r'[\x0b]'

        for value in row:
            if isinstance(value, str):
                yield re.sub(ILLEGAL_CHARACTERS, '', value)
            else:
                yield value

    def do_export(self):
        self.worksheet.append([force_text(h) for h in self.get_header()])

        for (i, instance) in enumerate(self.queryset.iterator()):
            try:
                self.worksheet.append(self.get_row(i, instance))
            except IllegalCharacterError:
                self.worksheet.append(self.clean_illegal_characters(self.get_row(i, instance)))

        self.add_style()

        archive = zipfile.ZipFile(
            self.output_file,
            mode='w',
            compression=zipfile.ZIP_DEFLATED,
            allowZip64=True
        )
        writer = ExcelWriter(self.workbook, archive)
        writer.write_data()
        # Finalize the zip archive
        archive.close()
